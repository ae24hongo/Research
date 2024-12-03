import openpyxl
import networkx as nx
import matplotlib.pyplot as plt

def load_excel_data(file_path, sheet_name, column_index=None):
    """
    Excelファイルからデータを読み込む

    Args:
        file_path (str): Excelファイルのパス
        sheet_name (str): 読み込むシート名
        column_index (int, optional): 特定の列のみ読み込む場合のインデックス

    Returns:
        list: 読み込んだデータのリスト
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    value = []
    if column_index is not None:
        # 特定の列のみ読み込む
        for row in ws.iter_rows(min_row=1):
            if len(row) > column_index:
                value.append(row[column_index].value)
    else:
        # 全行全列読み込み
        for row in ws.iter_rows(min_row=1):
            value_list = [c.value for c in row]
            value.append(value_list)
    
    return value

def get_company_name(code, reference_data):
    """
    企業コードから企業名を取得

    Args:
        code (str): 企業コード
        reference_data (list): 参照用のデータ

    Returns:
        str: 企業名（見つからない場合は元のコード）
    """
    for ref_row in reference_data:
        if str(code) == str(ref_row[0]):
            return ref_row[1]
    return code

def process_causality_data(causality_data, reference_data):
    """
    因果関係データを処理し、参照データと紐付ける

    Args:
        causality_data (list): 因果関係のデータ
        reference_data (list): 参照用のデータ

    Returns:
        list: 処理後のデータ
    """
    processed_data = []
    for cell_value in causality_data:
        if cell_value is None:
            continue
        
        processed_data.append(get_company_name(cell_value, reference_data))
    
    return processed_data

def create_network_graph(main_node, connected_nodes):
    """
    ネットワークグラフを作成

    Args:
        main_node (str): 中心となるノード
        connected_nodes (list): メインノードに接続されるノード

    Returns:
        nx.DiGraph: 有向グラフ
    """
    G = nx.DiGraph()
    G.add_node(main_node)
    
    edges = [(main_node, node) for node in connected_nodes]
    G.add_edges_from(edges)
    
    return G

def plot_network_graph(G, title='Causality Network Graph'):
    """
    ネットワークグラフをプロット

    Args:
        G (nx.DiGraph): プロットする有向グラフ
        title (str, optional): グラフのタイトル. Defaults to 'Causality Network Graph'.
    """
    plt.figure(figsize=(28, 21))
    
    pos = nx.spring_layout(G, k=2)  # ノード間の間隔を広げる
    cmap = plt.get_cmap("tab20")
    node_color = [cmap(i) for i in range(G.number_of_nodes())]
    
    nx.draw(
        G, 
        pos=pos,
        with_labels=True, 
        font_family='MS Gothic', 
        font_size=28, 
        node_size=5000, 
        width=5, 
        arrowsize=100,
        edge_color='skyblue', 
        node_color=node_color
    )
    
    plt.title(title, fontsize=32)
    plt.tight_layout()
    plt.show()

def main():
    # データ読み込み
    reference_data = load_excel_data("nikkei.xlsx", "Sheet1")
    
    # 因果関係データの読み込み（最初の列を読み込む）
    causality_data = load_excel_data("causality.xlsx", "Sheet1", column_index=0)
    
    # データから Noneを除外
    causality_data = [data for data in causality_data if data is not None]
    
    # メインノードの企業名を取得
    main_node_name = get_company_name(causality_data[0], reference_data)
    
    # データ処理
    processed_data = process_causality_data(causality_data[1:], reference_data)
    
    # ネットワークグラフの作成とプロット
    G = create_network_graph(main_node_name, processed_data)
    plot_network_graph(G, title=f'{main_node_name} - Causality Network')

if __name__ == "__main__":
    main()
