import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from statsmodels.tsa.vector_ar.var_model import VAR

def get_column_header(sheet, column_index):
    """
    指定されたシートと列インデックスから列ヘッダーを取得

    Args:
        sheet (openpyxl.worksheet.Worksheet): 対象のワークシート
        column_index (int): 列インデックス

    Returns:
        str: 列ヘッダーの値
    """
    column_letter = openpyxl.utils.get_column_letter(column_index)
    return sheet[column_letter + '1'].value

def load_excel_data(file_path):
    """
    Excelファイルからデータを読み込む

    Args:
        file_path (str): 読み込むExcelファイルのパス

    Returns:
        pd.DataFrame: 読み込んだデータのDataFrame
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active
    num_cols = sheet.max_column
    
    data = {}
    for col_num in range(2, num_cols + 1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        column_key = f'Column{col_num}'
        data[column_key] = [cell.value for cell in sheet[column_letter][1:]]
    
    return pd.DataFrame(data), sheet, num_cols

def perform_granger_causality_test(column_data, max_lags=10):
    """
    グレンジャー因果性検定を実行

    Args:
        column_data (np.ndarray): 分析対象の列データ
        max_lags (int, optional): 最大ラグ数. Defaults to 10.

    Returns:
        tuple: モデル結果、因果性検定結果（可能な場合）
    """
    model = VAR(column_data)
    lag = model.select_order(max_lags).selected_orders
    
    if lag['aic'] == 0:
        return None, None
    
    results = model.fit(lag['aic'])
    test_results0to1 = results.test_causality(causing=0, caused=1)
    
    return results, test_results0to1

def plot_impulse_response(var_results, period=10):
    """
    インパルス応答関数をプロット

    Args:
        var_results: VARモデルの結果
        period (int, optional): インパルス応答関数の期間. Defaults to 10.
    """
    irf = var_results.irf(period)
    irf.plot(orth=True)
    plt.show()

def analyze_causality(file_path, output_file):
    """
    グレンジャー因果性分析の全体処理

    Args:
        file_path (str): 入力Excelファイルのパス
        output_file (str): 出力Excelファイルのパス
    """
    # Excelデータ読み込み
    df, sheet, num_cols = load_excel_data(file_path)
    y = df.to_numpy()
    
    # 出力用Excelファイルの準備
    write_wb = load_workbook(output_file)
    write_ws = write_wb["Sheet1"]
    
    # 列ヘッダーの書き込み
    for col_num in range(2, num_cols + 1):
        column_header = get_column_header(sheet, col_num)
        write_ws.cell(row=1, column=col_num-1, value=column_header)
    
    # グレンジャー因果性分析
    for column_i in range(num_cols - 1):
        column_i_header = get_column_header(sheet, column_i + 2)
        
        for column_j in range(num_cols - 1):
            if column_i == column_j:
                continue
            
            column_data = y[:, [column_i, column_j]]
            results, test_results0to1 = perform_granger_causality_test(column_data)
            
            if results is None:
                continue
            
            column_j_header = get_column_header(sheet, column_j + 2)
            
            if test_results0to1.pvalue < 0.05:
                write_ws.cell(row=column_j + 2, column=column_i + 1, value=column_j_header)
                print(f"{column_i_header} to {column_j_header}")
                
                # インパルス応答関数のプロット
                plot_impulse_response(results)
    
    # 結果の保存
    write_wb.save(output_file)

def main():
    input_file = 'datalist_diff.xlsx'
    output_file = 'causality.xlsx'
    analyze_causality(input_file, output_file)

if __name__ == "__main__":
    main()
